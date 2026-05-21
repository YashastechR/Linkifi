"""
=============================================================
TASK 3: Excel Integration & Automation
=============================================================
Topics: Read Excel (multi-sheet), Write Excel, ExcelWriter,
        Format output, Automate reports

WHERE TO RUN:
  - Terminal:  python 03_excel_automation.py
  - Output files saved to: output/ folder
  - Open them in Microsoft Excel / LibreOffice / Google Sheets
=============================================================
"""

import pandas as pd
import numpy as np
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

print("=" * 60)
print("TASK 3: EXCEL AUTOMATION")
print("=" * 60)

# ─────────────────────────────────────────────────────────────
# SAMPLE DATA
# ─────────────────────────────────────────────────────────────
np.random.seed(42)

months = pd.date_range("2024-01-01", periods=12, freq="M")
month_labels = months.strftime("%b-%Y")

sales_df = pd.DataFrame({
    "Month":       month_labels,
    "North":       np.random.randint(50000, 150000, 12),
    "South":       np.random.randint(40000, 130000, 12),
    "East":        np.random.randint(45000, 120000, 12),
    "West":        np.random.randint(55000, 140000, 12),
})
sales_df["Total"] = sales_df[["North", "South", "East", "West"]].sum(axis=1)

employees_df = pd.DataFrame({
    "EmpID":      range(101, 116),
    "Name":       ["Aarav", "Bhavya", "Chetan", "Deepa", "Eshan",
                   "Farida", "Gaurav", "Heena", "Ishan", "Jyoti",
                   "Karan", "Lakshmi", "Mohan", "Nisha", "Om"],
    "Department": ["Sales", "IT", "HR", "Sales", "IT",
                   "HR", "Finance", "Sales", "IT", "Finance",
                   "HR", "Sales", "IT", "Finance", "Sales"],
    "Salary":     np.random.randint(35000, 120000, 15),
    "Bonus":      np.random.randint(5000, 25000, 15),
    "Rating":     np.round(np.random.uniform(3.0, 5.0, 15), 1),
})
employees_df["Total_Pay"] = employees_df["Salary"] + employees_df["Bonus"]

inventory_df = pd.DataFrame({
    "Product":   ["Laptop", "Phone", "Tablet", "Earbuds", "Watch",
                  "Camera", "Speaker", "Keyboard", "Mouse", "Monitor"],
    "Category":  ["Electronics"] * 5 + ["Accessories"] * 5,
    "Stock":     np.random.randint(5, 100, 10),
    "Price":     [65000, 25000, 35000, 2000, 15000,
                  45000, 8000, 3500, 1200, 22000],
    "Reorder_At": [10, 15, 10, 20, 8, 5, 12, 25, 30, 7],
})
inventory_df["Value"] = inventory_df["Stock"] * inventory_df["Price"]
inventory_df["Needs_Reorder"] = inventory_df["Stock"] <= inventory_df["Reorder_At"]

# ─────────────────────────────────────────────────────────────
# 1. WRITE TO EXCEL — single sheet
# ─────────────────────────────────────────────────────────────
print("\n─── 1. Write single-sheet Excel ───")

sales_path = OUTPUT_DIR / "sales_simple.xlsx"
sales_df.to_excel(sales_path, index=False, sheet_name="Sales_2024")
print(f"  Saved: {sales_path}")

# ─────────────────────────────────────────────────────────────
# 2. WRITE MULTIPLE SHEETS in one workbook
# ─────────────────────────────────────────────────────────────
print("\n─── 2. Write multiple sheets ───")

multi_path = OUTPUT_DIR / "company_data.xlsx"
with pd.ExcelWriter(multi_path, engine="openpyxl") as writer:
    sales_df.to_excel(writer,     sheet_name="Sales",     index=False)
    employees_df.to_excel(writer, sheet_name="Employees", index=False)
    inventory_df.to_excel(writer, sheet_name="Inventory", index=False)
print(f"  Saved: {multi_path}  (3 sheets)")

# ─────────────────────────────────────────────────────────────
# 3. READ BACK MULTIPLE SHEETS
# ─────────────────────────────────────────────────────────────
print("\n─── 3. Read Excel (multiple sheets) ───")

# Read all sheets at once → returns dict of DataFrames
all_sheets = pd.read_excel(multi_path, sheet_name=None)
for sheet_name, df in all_sheets.items():
    print(f"  Sheet '{sheet_name}': {df.shape[0]} rows × {df.shape[1]} cols")

# Read specific sheet
sales_back = pd.read_excel(multi_path, sheet_name="Sales")
print(f"\n  Sales sheet head:\n{sales_back.head(3)}")

# ─────────────────────────────────────────────────────────────
# 4. FORMATTED EXCEL REPORT with openpyxl styling
# ─────────────────────────────────────────────────────────────
print("\n─── 4. Formatted Excel Report ───")

from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

formatted_path = OUTPUT_DIR / "formatted_report.xlsx"

with pd.ExcelWriter(formatted_path, engine="openpyxl") as writer:

    # ── Sheet 1: Sales Report ──────────────────────────────
    sales_df.to_excel(writer, sheet_name="Sales_Report", index=False)
    ws = writer.sheets["Sales_Report"]

    # Header style
    header_fill  = PatternFill("solid", fgColor="1F4E79")   # dark blue
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[1]:          # row 1 = headers
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border

    # Alternate row shading
    light_blue = PatternFill("solid", fgColor="D6E4F0")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = light_blue if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = center_align

    # Currency format for numeric columns (B–F)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
        for cell in row:
            cell.number_format = '₹#,##0'

    # Auto-width columns
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col_letter].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # Color scale on "Total" column (column F)
    ws.conditional_formatting.add(
        f"F2:F{len(sales_df)+1}",
        ColorScaleRule(start_color="FF0000", mid_color="FFFF00", end_color="00FF00",
                       start_type="min", mid_type="percentile", mid_value=50,
                       end_type="max")
    )

    # ── Sheet 2: Employee Report ───────────────────────────
    employees_df.to_excel(writer, sheet_name="Employee_Report", index=False)
    ws2 = writer.sheets["Employee_Report"]

    # Header styling
    green_fill = PatternFill("solid", fgColor="1A6B3C")
    for cell in ws2[1]:
        cell.fill = green_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = center_align

    # Salary column → currency format
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=8):
        for cell in row:
            cell.number_format = '₹#,##0'

    for col in ws2.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws2.column_dimensions[col_letter].width = max_len + 4

    ws2.freeze_panes = "A2"

    # ── Sheet 3: Inventory Report ──────────────────────────
    inventory_df.to_excel(writer, sheet_name="Inventory", index=False)
    ws3 = writer.sheets["Inventory"]

    orange_fill = PatternFill("solid", fgColor="C55A11")
    red_fill    = PatternFill("solid", fgColor="FF0000")

    for cell in ws3[1]:
        cell.fill = orange_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = center_align

    # Highlight low-stock rows in red
    nr_col_idx = inventory_df.columns.get_loc("Needs_Reorder")  # 0-based
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        needs_reorder = row[nr_col_idx].value   # Needs_Reorder column
        if needs_reorder:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFD7D7")

    for col in ws3.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws3.column_dimensions[col_letter].width = max_len + 4

    ws3.freeze_panes = "A2"

print(f"  Saved: {formatted_path}  (3 styled sheets)")

# ─────────────────────────────────────────────────────────────
# 5. AUTOMATED SUMMARY REPORT with charts (openpyxl)
# ─────────────────────────────────────────────────────────────
print("\n─── 5. Automated Executive Summary Report ───")

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

report_path = OUTPUT_DIR / "executive_summary.xlsx"
wb = Workbook()

# ── Summary sheet ──────────────────────────────────────────
ws_sum = wb.active
ws_sum.title = "Summary"

# Title
ws_sum["A1"] = "ANNUAL SALES EXECUTIVE SUMMARY – 2024"
ws_sum["A1"].font      = Font(size=16, bold=True, color="1F4E79")
ws_sum["A1"].alignment = Alignment(horizontal="center")
ws_sum.merge_cells("A1:G1")

# KPI block
kpis = [
    ("Total Revenue",    f"₹{sales_df['Total'].sum():,.0f}"),
    ("Best Month",       sales_df.loc[sales_df['Total'].idxmax(), 'Month']),
    ("Best Region",      sales_df[["North","South","East","West"]].sum().idxmax()),
    ("Avg Monthly",      f"₹{sales_df['Total'].mean():,.0f}"),
    ("Peak Sales",       f"₹{sales_df['Total'].max():,.0f}"),
    ("Total Employees",  str(len(employees_df))),
]
ws_sum["A3"] = "KEY METRICS"
ws_sum["A3"].font = Font(bold=True, size=12, color="FFFFFF")
ws_sum["A3"].fill = PatternFill("solid", fgColor="1F4E79")
ws_sum.merge_cells("A3:B3")

for i, (label, value) in enumerate(kpis, start=4):
    ws_sum[f"A{i}"] = label
    ws_sum[f"B{i}"] = value
    ws_sum[f"A{i}"].font = Font(bold=True)
    ws_sum[f"A{i}"].fill = PatternFill("solid", fgColor="D6E4F0")
    ws_sum[f"B{i}"].alignment = Alignment(horizontal="right")

# ── Sales Data sheet for chart ──────────────────────────────
ws_data = wb.create_sheet("Sales_Data")
headers = ["Month", "North", "South", "East", "West", "Total"]
ws_data.append(headers)
for _, row_data in sales_df.iterrows():
    ws_data.append(list(row_data))

# Header styling
for cell in ws_data[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center")

# ── Bar Chart: Regional Sales ──────────────────────────────
bar_chart = BarChart()
bar_chart.type    = "col"
bar_chart.title   = "Monthly Regional Sales 2024"
bar_chart.y_axis.title = "Sales (₹)"
bar_chart.x_axis.title = "Month"
bar_chart.width   = 25
bar_chart.height  = 15

# Regions: columns 2–5 (North, South, East, West)
data_ref = Reference(ws_data, min_col=2, max_col=5,
                     min_row=1, max_row=13)
cats_ref = Reference(ws_data, min_col=1,
                     min_row=2, max_row=13)
bar_chart.add_data(data_ref, titles_from_data=True)
bar_chart.set_categories(cats_ref)
ws_data.add_chart(bar_chart, "A16")

# ── Line Chart: Total Trend ─────────────────────────────────
line_chart = LineChart()
line_chart.title          = "Total Monthly Sales Trend"
line_chart.y_axis.title   = "Total Sales (₹)"
line_chart.x_axis.title   = "Month"
line_chart.width  = 25
line_chart.height = 12

total_ref = Reference(ws_data, min_col=6, max_col=6,
                      min_row=1, max_row=13)
line_chart.add_data(total_ref, titles_from_data=True)
line_chart.set_categories(cats_ref)
ws_data.add_chart(line_chart, "A35")

# Column widths
for col in ws_data.columns:
    col_letter = get_column_letter(col[0].column)
    ws_data.column_dimensions[col_letter].width = 14

ws_sum.column_dimensions["A"].width = 22
ws_sum.column_dimensions["B"].width = 20

wb.save(report_path)
print(f"  Saved: {report_path}  (Summary + Charts)")

# ─────────────────────────────────────────────────────────────
# 6. READ EXCEL with specific options
# ─────────────────────────────────────────────────────────────
print("\n─── 6. Read Excel — advanced options ───")

# Skip rows, use specific columns, set dtype
df_read = pd.read_excel(
    formatted_path,
    sheet_name="Employee_Report",
    usecols=["Name", "Department", "Salary", "Rating"],
    dtype={"Salary": float}
)
print(f"  Read {df_read.shape[0]} rows from Employee_Report")
print(df_read.head())

dept_summary = df_read.groupby("Department")["Salary"].agg(["mean", "max", "count"]).round(0)
print("\n  Department summary:")
print(dept_summary)

print("\n✅  Task 3 complete — all Excel files saved to output/")
print("   Files created:")
for f in sorted(OUTPUT_DIR.glob("*.xlsx")):
    print(f"     • {f.name}")